"""
LRZ Social Media ETL Pipeline
==============================
Reads native exports from LinkedIn, Instagram, and ArtStation,
normalises to monthly aggregates, and writes directly to the
LRZ_Social_Media_Dashboard.xlsx Raw Data sheet.

Usage:
    python lrz_etl.py [--xlsx PATH] [--data-dir PATH] [--dry-run]

Expected files in data-dir/
    linkedin_updates.csv   — LinkedIn Analytics > Updates > Export
    linkedin_followers.csv — LinkedIn Analytics > Followers > Export
    instagram_insights.csv — Meta Business Suite > Insights > Export
    artstation_manual.csv  — Manual template (see template in data-dir/)
"""

import argparse
import os
import sys
import json
import re
import textwrap
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

import pandas as pd
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Colour palette (must match dashboard) ───────────────────────────────────
C = {
    'dark_bg':    'FF1A1A2E', 'mid_bg':   'FF16213E',
    'panel':      'FF0F3460', 'row_alt':  'FF1E2A4A',
    'red':        'FFE94560', 'gold':     'FFFDCB6E',
    'teal':       'FF00B4D8', 'green':    'FF00C48C',
    'white':      'FFFFFFFF', 'muted':    'FF8A8FA3',
    'linkedin':   'FF0077B5', 'instagram':'FFE1306C',
    'twitter':    'FF1DA1F2', 'artstation':'FF13AFF0',
    'facebook':   'FF1877F2', 'header_bg':'FF0F3460',
    'subhdr_bg':  'FF16213E', 'orange':   'FFFF9A3C',
}

def _fill(hex_color): return PatternFill('solid', start_color=hex_color[2:], end_color=hex_color[2:])
def _font(bold=False, size=10, color='FFFFFFFF', italic=False):
    return Font(bold=bold, size=size, color=color[2:], italic=italic, name='Calibri')
def _border():
    s = Side(style='thin', color='404060')
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h='center', wrap=False):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

def style(cell, value=None, bg=None, fg=None, bold=False, size=10,
          h='center', wrap=False, fmt=None, border=True, italic=False):
    if value is not None: cell.value = value
    if bg: cell.fill = _fill(bg)
    cell.font = _font(bold=bold, size=size, color=fg or C['white'], italic=italic)
    cell.alignment = _align(h=h, wrap=wrap)
    if border: cell.border = _border()
    if fmt: cell.number_format = fmt
    return cell


# ─── PARSERS ──────────────────────────────────────────────────────────────────

class ParseResult:
    """Container for normalised monthly data from one platform."""
    def __init__(self, platform: str):
        self.platform = platform
        self.rows: dict[str, dict] = {}   # key = 'YYYY-MM'
        self.errors: list[str] = []
        self.source_files: list[str] = []

    def add_month(self, month_key: str, **kwargs):
        if month_key not in self.rows:
            self.rows[month_key] = {
                'followers': None, 'impressions': 0, 'engagements': 0,
                'posts': 0, 'likes': 0, 'comments': 0, 'shares': 0, 'reach': 0
            }
        for k, v in kwargs.items():
            if v is not None:
                if k in ('impressions','engagements','posts','likes','comments','shares','reach'):
                    self.rows[month_key][k] = (self.rows[month_key].get(k) or 0) + v
                else:
                    self.rows[month_key][k] = v

    def sorted_months(self):
        return sorted(self.rows.keys())


def _to_month_key(dt) -> str:
    return dt.strftime('%Y-%m')

def _parse_date_flexible(s: str) -> date | None:
    s = str(s).strip()
    for fmt in ('%m/%d/%Y','%Y-%m-%d','%d/%m/%Y','%b %Y','%B %Y','%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── LinkedIn ──────────────────────────────────────────────────────────────────

def parse_linkedin(updates_path: str | None, followers_path: str | None) -> ParseResult:
    result = ParseResult('linkedin')

    # ── Followers file ──
    if followers_path and os.path.exists(followers_path):
        result.source_files.append(followers_path)
        try:
            df = pd.read_csv(followers_path)
            df.columns = [c.strip().lower() for c in df.columns]
            date_col = next((c for c in df.columns if 'date' in c), None)
            fol_col  = next((c for c in df.columns if 'total' in c and 'follow' in c), None)
            if date_col and fol_col:
                for _, row in df.iterrows():
                    d = _parse_date_flexible(row[date_col])
                    if d:
                        mk = _to_month_key(d)
                        try: result.add_month(mk, followers=int(float(row[fol_col])))
                        except (ValueError, TypeError): pass
            else:
                result.errors.append(f"LinkedIn followers: could not find Date/Total followers columns. Found: {list(df.columns)}")
        except Exception as e:
            result.errors.append(f"LinkedIn followers parse error: {e}")

    # ── Updates file ──
    if updates_path and os.path.exists(updates_path):
        result.source_files.append(updates_path)
        try:
            df = pd.read_csv(updates_path)
            df.columns = [c.strip().lower() for c in df.columns]
            date_col = next((c for c in df.columns if 'date' in c), None)

            def _col(keywords):
                for c in df.columns:
                    if all(k in c for k in keywords): return c
                return None

            imp_col     = _col(['impressions','organic']) or _col(['impressions'])
            react_col   = _col(['reactions','organic'])   or _col(['reactions'])
            comment_col = _col(['comments','organic'])    or _col(['comments'])
            share_col   = _col(['shares','organic'])      or _col(['shares'])

            if not date_col:
                result.errors.append("LinkedIn updates: no Date column found")
            else:
                monthly = defaultdict(lambda: defaultdict(float))
                monthly_posts = defaultdict(int)
                for _, row in df.iterrows():
                    d = _parse_date_flexible(row[date_col])
                    if not d: continue
                    mk = _to_month_key(d)
                    monthly_posts[mk] += 1
                    def _val(col):
                        if col is None: return 0
                        try: return float(row.get(col, 0) or 0)
                        except (ValueError, TypeError): return 0
                    monthly[mk]['impressions']  += _val(imp_col)
                    monthly[mk]['reactions']    += _val(react_col)
                    monthly[mk]['comments']     += _val(comment_col)
                    monthly[mk]['shares']       += _val(share_col)

                for mk, agg in monthly.items():
                    engagements = int(agg['reactions'] + agg['comments'] + agg['shares'])
                    result.add_month(mk,
                        impressions=int(agg['impressions']),
                        engagements=engagements,
                        likes=int(agg['reactions']),
                        comments=int(agg['comments']),
                        shares=int(agg['shares']),
                        posts=monthly_posts[mk]
                    )
        except Exception as e:
            result.errors.append(f"LinkedIn updates parse error: {e}")

    return result


# ── Instagram ─────────────────────────────────────────────────────────────────

def parse_instagram(insights_path: str | None) -> ParseResult:
    result = ParseResult('instagram')
    if not insights_path or not os.path.exists(insights_path):
        result.errors.append("Instagram insights file not found")
        return result

    result.source_files.append(insights_path)
    try:
        df = pd.read_csv(insights_path, comment='#')
        df.columns = [c.strip().lower() for c in df.columns]

        date_col    = next((c for c in df.columns if c in ('period','date','week')), None)
        reach_col   = next((c for c in df.columns if 'reach' in c or 'accounts reached' in c), None)
        imp_col     = next((c for c in df.columns if 'impression' in c), None)
        likes_col   = next((c for c in df.columns if 'like' in c), None)
        comments_col= next((c for c in df.columns if 'comment' in c), None)
        shares_col  = next((c for c in df.columns if 'share' in c), None)
        saves_col   = next((c for c in df.columns if 'save' in c), None)
        follows_col = next((c for c in df.columns if 'follow' in c and 'new' in c), None) or \
                      next((c for c in df.columns if 'follow' in c), None)

        if not date_col:
            result.errors.append(f"Instagram: no date column. Columns: {list(df.columns)}")
            return result

        monthly = defaultdict(lambda: defaultdict(float))
        monthly_posts = defaultdict(int)
        monthly_follows = defaultdict(int)

        for _, row in df.iterrows():
            d = _parse_date_flexible(row[date_col])
            if not d: continue
            mk = _to_month_key(d)
            monthly_posts[mk] += 1

            def _v(col):
                if col is None: return 0
                try: return float(row.get(col, 0) or 0)
                except (ValueError, TypeError): return 0

            monthly[mk]['impressions'] += _v(imp_col)
            monthly[mk]['reach']       += _v(reach_col)
            monthly[mk]['likes']       += _v(likes_col)
            monthly[mk]['comments']    += _v(comments_col)
            monthly[mk]['shares']      += _v(shares_col)
            monthly[mk]['saves']       += _v(saves_col)
            monthly_follows[mk]        += int(_v(follows_col))

        # Reconstruct followers from accumulated follows
        known_start = 2350
        sorted_months_ig = sorted(monthly.keys())
        running_followers = known_start
        follower_map = {}
        for mk in sorted_months_ig:
            running_followers += monthly_follows[mk]
            follower_map[mk] = running_followers

        for mk, agg in monthly.items():
            engagements = int(agg['likes'] + agg['comments'] + agg['shares'] + agg['saves'])
            result.add_month(mk,
                followers=follower_map.get(mk),
                impressions=int(agg['impressions']),
                reach=int(agg['reach']),
                engagements=engagements,
                likes=int(agg['likes']),
                comments=int(agg['comments']),
                shares=int(agg['shares']),
                posts=monthly_posts[mk]
            )
    except Exception as e:
        result.errors.append(f"Instagram parse error: {e}")

    return result


# ── ArtStation ────────────────────────────────────────────────────────────────

def parse_artstation(manual_path: str | None) -> ParseResult:
    result = ParseResult('artstation')
    if not manual_path or not os.path.exists(manual_path):
        result.errors.append("ArtStation manual CSV not found")
        return result

    result.source_files.append(manual_path)
    try:
        df = pd.read_csv(manual_path, comment='#')
        df.columns = [c.strip().lower() for c in df.columns]

        month_col   = next((c for c in df.columns if 'month' in c), None)
        fol_col     = next((c for c in df.columns if 'follower' in c), None)
        views_col   = next((c for c in df.columns if 'view' in c), None)
        likes_col   = next((c for c in df.columns if 'like' in c), None)
        comment_col = next((c for c in df.columns if 'comment' in c), None)
        posts_col   = next((c for c in df.columns if 'post' in c or 'project' in c), None)

        if not month_col:
            result.errors.append(f"ArtStation: no Month column. Columns: {list(df.columns)}")
            return result

        for _, row in df.iterrows():
            d = _parse_date_flexible(row[month_col])
            if not d: continue
            mk = _to_month_key(d)

            def _v(col):
                if col is None: return None
                try: return int(float(row.get(col, 0) or 0))
                except (ValueError, TypeError): return None

            fol   = _v(fol_col)
            views = _v(views_col)
            likes = _v(likes_col)
            cmts  = _v(comment_col)
            posts = _v(posts_col)
            eng = ((likes or 0) + (cmts or 0))
            result.add_month(mk,
                followers=fol,
                impressions=views or 0,
                engagements=eng,
                likes=likes or 0,
                comments=cmts or 0,
                posts=posts or 0
            )
    except Exception as e:
        result.errors.append(f"ArtStation parse error: {e}")

    return result


# ─── XLSX WRITER ──────────────────────────────────────────────────────────────

PLATFORM_META = {
    'linkedin':   ('LINKEDIN',   C['linkedin']),
    'instagram':  ('INSTAGRAM',  C['instagram']),
    'twitter':    ('X / TWITTER',C['twitter']),
    'artstation': ('ARTSTATION', C['artstation']),
    'facebook':   ('FACEBOOK',   C['facebook']),
}

RAW_DATA_COLS = [
    ('Month',        '@',     14),
    ('Followers',    '#,##0', 12),
    ('MoM Chg',      '+#,##0;-#,##0;-', 12),
    ('MoM %',        '0.0%',  12),
    ('Posts',        '0',     10),
    ('Impressions',  '#,##0', 13),
    ('Engagements',  '#,##0', 12),
    ('Eng Rate %',   '0.0%',  12),
    ('Avg Likes',    '#,##0', 12),
    ('Comments',     '#,##0', 12),
    ('Shares',       '#,##0', 11),
    ('Reach',        '#,##0', 11),
    ('Data Source',  '@',     20),
]


def _write_platform_block(ws, start_row: int, platform_key: str, result: ParseResult,
                          col_offset: int = 2) -> int:
    """Write one platform's monthly data block. Returns the next available row."""
    plat_label, plat_color = PLATFORM_META.get(platform_key, (platform_key.upper(), C['panel']))
    source_tag = ', '.join(Path(f).name for f in result.source_files) or '— manual/estimated'

    # Platform header
    r = start_row
    ws.row_dimensions[r].height = 24
    last_col = col_offset + len(RAW_DATA_COLS) - 1
    ws.merge_cells(start_row=r, start_column=col_offset, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=col_offset)
    c.value = f"● {plat_label}  —  source: {source_tag}"
    c.fill = _fill(plat_color)
    c.font = _font(bold=True, size=11)
    c.alignment = _align(h='left')
    c.border = _border()
    r += 1

    # Column headers
    ws.row_dimensions[r].height = 20
    for i, (hdr, _, _) in enumerate(RAW_DATA_COLS):
        c = ws.cell(row=r, column=col_offset+i)
        style(c, value=hdr, bg=C['subhdr_bg'], fg=C['gold'], bold=True, size=9)
    r += 1

    months = result.sorted_months()
    data_start_row = r

    for idx, mk in enumerate(months):
        row_data = result.rows[mk]
        ws.row_dimensions[r].height = 18
        bg = C['row_alt'] if idx % 2 == 0 else C['mid_bg']

        col = col_offset
        fol = row_data.get('followers')
        imp = row_data.get('impressions', 0)
        eng = row_data.get('engagements', 0)
        posts = row_data.get('posts', 0)
        likes = row_data.get('likes', 0)
        cmts  = row_data.get('comments', 0)
        shares= row_data.get('shares', 0)
        reach = row_data.get('reach', 0)
        avg_likes = int(likes * 0.75) if likes else 0

        # Month label
        try:
            d = datetime.strptime(mk, '%Y-%m')
            month_label = d.strftime('%b %Y')
        except ValueError:
            month_label = mk

        col_fol_letter  = get_column_letter(col_offset + 1)
        col_imp_letter  = get_column_letter(col_offset + 5)
        col_eng_letter  = get_column_letter(col_offset + 6)

        values = [
            month_label,
            fol if fol is not None else '',
            f'={col_fol_letter}{r}-{col_fol_letter}{r-1}' if idx > 0 and fol is not None else '',
            f'=IFERROR(({col_fol_letter}{r}-{col_fol_letter}{r-1})/{col_fol_letter}{r-1},0)' if idx > 0 and fol is not None else '',
            posts,
            imp,
            eng,
            f'=IFERROR({col_eng_letter}{r}/{col_imp_letter}{r},0)',
            avg_likes,
            cmts,
            shares,
            reach,
            'Live export' if result.source_files else 'Estimated'
        ]

        for i, (val, (_, fmt, _)) in enumerate(zip(values, RAW_DATA_COLS)):
            c = ws.cell(row=r, column=col_offset + i)
            style(c, value=val, bg=bg, fg=C['white'], size=9, fmt=fmt)
            # MoM % colour
            if i == 3 and idx > 0:
                c.font = _font(size=9, color=C['green'])
            # Eng Rate colour
            if i == 7:
                c.font = _font(size=9, color=C['teal'])

        r += 1

    return r + 1  # one spacer row


def write_import_zone_sheet(wb, results_by_platform: dict, import_summary: dict):
    """Add/overwrite the ⬇ Import Zone sheet with instructions + status."""
    sheet_name = "⬇ Import Zone"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 1)
    ws.sheet_properties.tabColor = "FDCB6E"
    ws.sheet_view.showGridLines = False

    for r in range(1, 80):
        for c in range(1, 20):
            ws.cell(row=r, column=c).fill = _fill(C['dark_bg'])

    for i, w in enumerate([2,18,12,50,20,2], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells('B1:E1')
    ws.row_dimensions[1].height = 35
    style(ws['B1'], value="🧟 LRZ — DATA IMPORT STATUS & INSTRUCTIONS",
          bg=C['red'], bold=True, size=14)

    # Last run timestamp
    ws.merge_cells('B2:E2')
    ws.row_dimensions[2].height = 18
    style(ws['B2'],
          value=f"Last ETL run: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Raw Data sheet updated automatically",
          bg=C['mid_bg'], fg=C['gold'], size=9, italic=True)

    # Status table
    row = 4
    ws.merge_cells(f'B{row}:E{row}')
    ws.row_dimensions[row].height = 22
    style(ws.cell(row=row, column=2), value="◆  IMPORT STATUS PER PLATFORM",
          bg=C['panel'], bold=True, size=11, fg=C['gold'], h='left')
    row += 1

    hdrs = ['Platform', 'Status', 'Months Imported', 'Source Files']
    ws.row_dimensions[row].height = 20
    for i, h in enumerate(hdrs, 2):
        style(ws.cell(row=row, column=i), value=h,
              bg=C['subhdr_bg'], bold=True, size=9, fg=C['gold'])
    row += 1

    for i, (k, res) in enumerate(results_by_platform.items()):
        _, plat_color = PLATFORM_META.get(k, (k, C['panel']))
        bg = C['row_alt'] if i % 2 == 0 else C['mid_bg']
        status = '✅ Imported' if res.source_files else ('⚠ Errors' if res.errors else '— No file')
        status_color = C['green'] if res.source_files else (C['orange'] if res.errors else C['muted'])
        ws.row_dimensions[row].height = 18

        style(ws.cell(row=row, column=2), value=PLATFORM_META.get(k,(k,''))[0],
              bg=plat_color, bold=True, size=9)
        style(ws.cell(row=row, column=3), value=status,
              bg=bg, fg=status_color, bold=True, size=9)
        style(ws.cell(row=row, column=4), value=str(len(res.rows)) + ' months',
              bg=bg, size=9)
        src = ', '.join(Path(f).name for f in res.source_files) or '— '
        style(ws.cell(row=row, column=5), value=src, bg=bg, size=9, h='left')
        row += 1

    if any(r.errors for r in results_by_platform.values()):
        row += 1
        ws.merge_cells(f'B{row}:E{row}')
        ws.row_dimensions[row].height = 22
        style(ws.cell(row=row, column=2), value="◆  ERRORS & WARNINGS",
              bg=C['red'], bold=True, size=11, fg=C['white'], h='left')
        row += 1
        for k, res in results_by_platform.items():
            for err in res.errors:
                ws.merge_cells(f'B{row}:E{row}')
                ws.row_dimensions[row].height = 28
                style(ws.cell(row=row, column=2),
                      value=f"[{k.upper()}] {err}",
                      bg=C['mid_bg'], fg=C['orange'], size=9, h='left', wrap=True)
                row += 1

    # Instructions
    row += 1
    ws.merge_cells(f'B{row}:E{row}')
    ws.row_dimensions[row].height = 22
    style(ws.cell(row=row, column=2), value="◆  HOW TO EXPORT FROM EACH PLATFORM",
          bg=C['panel'], bold=True, size=11, fg=C['gold'], h='left')
    row += 1

    instructions = [
        ('LINKEDIN — Updates',
         '1. Go to linkedin.com/company/littleredzombies/admin/analytics/updates/\n'
         '2. Set date range > click Export > saves as linkedin_updates.csv\n'
         '3. Also export Followers: analytics/followers/ > Export > linkedin_followers.csv'),
        ('INSTAGRAM — Insights',
         '1. Go to business.facebook.com > Insights > Overview\n'
         '2. Set date range to custom (select all) > Export Data\n'
         '3. Choose "Account" data type, CSV format > saves as instagram_insights.csv'),
        ('ARTSTATION — Manual',
         '1. Go to artstation.com/littleredzombies > Dashboard\n'
         '2. Open artstation_manual.csv template\n'
         '3. Fill in monthly totals from your ArtStation dashboard stats panel\n'
         '   (Followers at end of month, total Views/Likes/Comments for each month)'),
        ('RUN THE PIPELINE',
         '1. Place all CSV files in the data/ folder next to lrz_etl.py\n'
         '2. Run: python lrz_etl.py\n'
         '3. The script auto-updates this XLSX in-place\n'
         '4. Check this sheet for any errors, then check 📋 Raw Data for results'),
    ]

    plat_instr_colors = [C['linkedin'], C['instagram'], C['artstation'], C['green']]
    for i, (title, body) in enumerate(instructions):
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f'B{row}:E{row}')
        style(ws.cell(row=row, column=2), value=title,
              bg=plat_instr_colors[i], bold=True, size=10, h='left')
        row += 1
        for line in body.strip().split('\n'):
            ws.row_dimensions[row].height = 18
            ws.merge_cells(f'B{row}:E{row}')
            style(ws.cell(row=row, column=2), value='  '+line,
                  bg=C['row_alt'] if row % 2 == 0 else C['mid_bg'],
                  fg=C['muted'], size=9, h='left')
            row += 1
        row += 1


def write_raw_data_sheet(wb, results_by_platform: dict):
    """Overwrite the 📋 Raw Data sheet with fresh imported data."""
    sheet_name = "📋 Raw Data"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 2)
    ws.sheet_properties.tabColor = "0077B5"
    ws.sheet_view.showGridLines = False

    for r in range(1, 200):
        for c in range(1, 18):
            ws.cell(row=r, column=c).fill = _fill(C['dark_bg'])

    # Column widths
    widths = [2] + [c[2] for c in RAW_DATA_COLS] + [2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells(f'B1:{get_column_letter(1+len(RAW_DATA_COLS))}1')
    ws.row_dimensions[1].height = 35
    style(ws['B1'],
          value="🧟 LITTLE RED ZOMBIES — MONTHLY SOCIAL MEDIA DATA (Live Import)",
          bg=C['red'], bold=True, size=14)

    # Data note
    ws.merge_cells(f'B2:{get_column_letter(1+len(RAW_DATA_COLS))}2')
    ws.row_dimensions[2].height = 18
    sources = {k: r.source_files for k, r in results_by_platform.items()}
    live = [k for k, f in sources.items() if f]
    est  = [k for k, f in sources.items() if not f]
    note = f"✅ Live data: {', '.join(live) or 'none'}   |   ⚠ Estimated: {', '.join(est) or 'none'}"
    style(ws.cell(row=2, column=2), value=note, bg=C['orange'],
          fg='FF1A1A2E', size=9, h='left')

    current_row = 4
    order = ['linkedin', 'instagram', 'twitter', 'artstation', 'facebook']
    for plat in order:
        res = results_by_platform.get(plat)
        if res is None or not res.rows:
            continue
        current_row = _write_platform_block(ws, current_row, plat, res)

    ws.freeze_panes = 'C6'


# ─── FALLBACK: keep estimated data for missing platforms ─────────────────────

def build_estimated_twitter() -> ParseResult:
    """Twitter has no export — return mock data."""
    res = ParseResult('twitter')
    res.errors.append("X/Twitter: no official CSV export available. Using estimated data.")
    followers = [940,960,978,994,1010,1024,1038,1050,1062,1072,1082,1092,
                 1100,1108,1118,1128,1140,1150,1162,1175,1190,1207,1220,1235,
                 1250,1265,1278,1290]
    months_keys = []
    d = date(2024,1,1)
    for _ in range(28):
        months_keys.append(_to_month_key(d))
        d += relativedelta(months=1)

    for mk, fol in zip(months_keys, followers):
        imp = fol * 1.4
        eng = int(imp * 0.012)
        res.add_month(mk, followers=fol,
                      impressions=int(imp), engagements=eng,
                      likes=int(eng*0.6), comments=int(eng*0.15),
                      shares=int(eng*0.25), posts=2)
    return res

def build_estimated_facebook() -> ParseResult:
    res = ParseResult('facebook')
    res.errors.append("Facebook: no CSV export run. Using estimated data.")
    followers = [1820,1850,1878,1904,1930,1952,1974,1996,2014,2030,2048,2064,
                 2082,2100,2116,2132,2148,2160,2172,2182,2192,2200,2206,2212,
                 2218,2224,2230,2238]
    months_keys = []
    d = date(2024,1,1)
    for _ in range(28):
        months_keys.append(_to_month_key(d))
        d += relativedelta(months=1)
    for mk, fol in zip(months_keys, followers):
        imp = int(fol * 1.0)
        eng = int(imp * 0.008)
        res.add_month(mk, followers=fol,
                      impressions=imp, engagements=eng,
                      likes=int(eng*0.7), comments=int(eng*0.1),
                      shares=int(eng*0.2), posts=2)
    return res


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def run(xlsx_path: str, data_dir: str, dry_run: bool = False):
    data_dir = Path(data_dir)
    print(f"\n{'='*60}")
    print(f"  LRZ Social Media ETL Pipeline")
    print(f"  Data dir : {data_dir.resolve()}")
    print(f"  XLSX     : {xlsx_path}")
    print(f"  Dry run  : {dry_run}")
    print(f"{'='*60}")

    # ── Parse all platforms ──
    print("\n[1/3] Parsing platform exports...")

    li_updates  = str(data_dir / 'linkedin_updates.csv')
    li_followers= str(data_dir / 'linkedin_followers.csv')
    ig_insights = str(data_dir / 'instagram_insights.csv')
    as_manual   = str(data_dir / 'artstation_manual.csv')

    results = {
        'linkedin':   parse_linkedin(li_updates, li_followers),
        'instagram':  parse_instagram(ig_insights),
        'artstation': parse_artstation(as_manual),
        'twitter':    build_estimated_twitter(),
        'facebook':   build_estimated_facebook(),
    }

    # Print summary
    for k, res in results.items():
        months_n = len(res.rows)
        src = ', '.join(Path(f).name for f in res.source_files) or '(estimated)'
        status = '✅' if res.source_files else '⚠'
        print(f"  {status} {k.upper():12s} — {months_n} months  source: {src}")
        for err in res.errors:
            print(f"     ⚠  {err}")

    if dry_run:
        print("\n[DRY RUN] No file written.")
        return results

    # ── Load workbook ──
    print(f"\n[2/3] Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)

    # ── Write sheets ──
    print("[3/3] Writing sheets...")
    write_import_zone_sheet(wb, results, {})
    write_raw_data_sheet(wb, results)

    wb.save(xlsx_path)
    print(f"\n✅ Saved: {xlsx_path}")

    # ── JSON log ──
    log = {
        'run_at': datetime.now().isoformat(),
        'xlsx': str(xlsx_path),
        'platforms': {
            k: {
                'months': len(v.rows),
                'source_files': v.source_files,
                'errors': v.errors,
                'first_month': v.sorted_months()[0] if v.rows else None,
                'last_month':  v.sorted_months()[-1] if v.rows else None,
            }
            for k, v in results.items()
        }
    }
    log_path = Path(xlsx_path).parent / 'lrz_etl_log.json'
    log_path.write_text(json.dumps(log, indent=2))
    print(f"   Log   : {log_path}")
    print()
    return results


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='LRZ Social Media ETL Pipeline')
    parser.add_argument('--xlsx',     default='LRZ_Social_Media_Dashboard.xlsx',
                        help='Path to the dashboard XLSX file')
    parser.add_argument('--data-dir', default='data/',
                        help='Directory containing platform CSVs')
    parser.add_argument('--dry-run',  action='store_true',
                        help='Parse only; do not write to XLSX')
    args = parser.parse_args()
    run(args.xlsx, args.data_dir, args.dry_run)
